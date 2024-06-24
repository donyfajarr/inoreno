from django.shortcuts import render, redirect
from django.contrib.auth import login as auth_login, logout, authenticate
from django.contrib import messages
from .decorators import login
from openpyxl import load_workbook
from datetime import datetime, timedelta, date
from . import models
from django.core.exceptions import ObjectDoesNotExist
from django.utils.dateparse import parse_date
# Create your views here.
from django.db.models import Q
import json
from django.http import JsonResponse
import requests
import urllib.parse
from urllib.parse import unquote
import csv
from django.http import HttpResponse
import ast

@login
def index(request):
    first_name = request.user.first_name
    last_name = request.user.last_name
    user = first_name + " " + last_name

    # Total Ongoing Projects
    sum = models.project.objects.filter(assignee = request.user).exclude(Q(status=3)).count()
    
    # Mencari Ongoing Project
    closeproject = models.project.objects.filter(assignee = request.user, status=3).count()

    # Total Semua Projects
    rawproject = models.project.objects.filter(assignee = request.user).count()
    rawprojects = models.project.objects.filter(assignee = request.user)
    
    # Total Ongoing Task
    totaltask = models.task.objects.filter(assignee = request.user).exclude(Q(status=3)).count()
    
    # Total Semua Task
    rawtask = models.task.objects.filter(assignee = request.user).count()
    
    # Mencari Closest Due Date Task
    alltask = models.task.objects.filter(assignee = request.user).order_by('due_date').exclude(Q(due_date=None))
    opentaskraw = models.task.objects.filter(assignee = request.user, status=1, due_date__isnull=False).order_by("due_date")
    
    # Jumlah Task yang open, close, dan underreview
    opentask = models.task.objects.filter(assignee = request.user, status=1).count()
    closetask = models.task.objects.filter(assignee = request.user, status=3).count()
    rawunder = models.task.objects.filter(assignee = request.user, status=2)

    # Jumlah underreview task
    undertask = rawunder.count()

    # Mencari tanggal hari ini
    today = date.today()
    start_of_week = today - timedelta(days=today.weekday())

    # Dibawah ini data untuk grafik pada dashboard
    categories = []
    open_counts = []
    closed_counts = []
    for i in range(5):
        day = start_of_week + timedelta(days=i)
        categories.append(day.strftime("%a"))
        closed_tasks = models.task.objects.filter(assignee = request.user, status=3, due_date = day).count()
        open_tasks = models.task.objects.filter(assignee = request.user, status__in=[1,2], due_date=day).count()
        closed_counts.append(closed_tasks)
        open_counts.append(-open_tasks)
    series = [
        {
            "name": "Closed",
            "data": closed_counts
        },
        {
            "name": "Open",
            "data": open_counts
        }
    ]

    chartdata = {
    "categories" : categories,
    "series" : series
    }

    data = {
        'labels' : ['Open Task', 'Under Review Task', 'Closed Task'],
        'series' : [opentask, undertask, closetask]
    }
    totalproject = models.project.objects.filter(assignee=request.user).count()
    closeproject = totalproject - sum
    reviewproject = totalproject - sum - closeproject

    json_chartdata = json.dumps(chartdata)
    json_data = json.dumps(data)
    return render(request, 'index.html', {
        'user':user,
        'sum':sum,
        'totaltask':totaltask,
        'alltask' : alltask,
        'rawtask' : rawtask,
        'rawproject' : rawproject,
        'json_data' : json_data,
        'opentask' :opentask,
        'undertask' :undertask,
        'closetask' :closetask,
        'rawunder':rawunder,
        'opentaskraw' : opentaskraw,
        'rawprojects' : rawprojects,
        'closeproject' : closeproject,
        'reviewproject' : reviewproject,
        'json_chartdata' : json_chartdata
    
    })

# Ini function untuk menghandle input dari grafik
def get_project_data(request):
    
    if request.method == 'GET':
        
        # Mengembalikan project terpilih
        selected_project = request.GET.get('project_id')
        print(selected_project)
        tasks = models.task.objects.filter(id_project=selected_project)
    
        # Mencari data updatean
        closed_counts = []
        open_counts = []
        categories = []
        today = date.today()
        start_of_week = today - timedelta(days=today.weekday())
        for i in range(5):
            day = start_of_week + timedelta(days=i)
            categories.append(day.strftime("%a"))
            closed_tasks = tasks.filter(assignee = request.user, status=3, due_date = day).count()
            open_tasks = tasks.filter(assignee = request.user, status__in=[1,2], due_date=day).count()
            closed_counts.append(closed_tasks)
            open_counts.append(-open_tasks)

        # Mengembalikan degan json untuk diload di grafik
        series = [
        {
            "name": "Closed",
            "data": closed_counts
        },
        {
            "name": "Open",
            "data": open_counts
        }
    ]
        data = {
            'series' : series
        }
        return JsonResponse(data)

    return JsonResponse({'error': 'Invalid request'})

# Ini Function untuk Tambah Task
@login
def addissue(request, id):
    if request.method == "GET":
        getallproject = models.project.objects.filter(assignee=request.user)
        allstatus = models.status.objects.all()
        allpriority = models.priority.objects.all()
        id = int(id)
        return render(request, 'addissue.html', {
            'getallproject' : getallproject,
            'allstatus' : allstatus,
            'allpriority' : allpriority,
            'id' : id,
        })
    else:
        id = request.POST['id']
        subject = request.POST['subject']
        status = request.POST['status']
        priority = request.POST['priority']
        start_date = request.POST['start_date'] 
        due_date = request.POST['due_date']
        parent = request.POST['parent']
        assignee = request.user
        pic = request.POST['pic']
        getstatus = models.status.objects.get(id=status)
        getpriority = models.priority.objects.get(id=priority)
        getproject = models.project.objects.get(id=id)
        createissue = models.task.objects.create(
            subject = subject,
            status = getstatus,
            id_project = getproject,
            id_priority = getpriority,
            start_date = start_date if start_date else None,
            due_date = due_date if due_date else None,
            assignee = assignee,
            parent = parent
        )
        createpic = models.pic.objects.create(id_task = createissue, pic=pic)
        return redirect ('listproject')
    

@login
def listproject(request):
    listproject = models.project.objects.filter(assignee = request.user)
    return render(request, 'listproject.html', {
        'listproject': listproject,})

@login
def newproject(request):
    if request.method == "POST":
        subject = request.POST['name']
        desc = request.POST['description']
        get = models.status.objects.get(id=1)
        newproject = models.project(
            subject = subject,
            desc = desc,
            assignee = request.user,
            status = get
        )
        newproject.save()
        return redirect('confirmation', id=newproject.id)
    return render(request, 'newproject.html')

# Ini function untuk mengubah settings dari baris dan kolom pada template excel
@login
def settings(request):
    user = request.user

    if request.method == "GET":

        try:
            get = models.settings.objects.get(user=user)
            
            dict = {
                'start_row' : get.start_row,
                'gannt_start_column' : get.gannt_start_column,
                'week_number_row' : get.week_number_row,
                'start_column' : get.start_column,
                'end_column' : get.end_column,
                'pic_column' : get.pic_column,
                'email_column' : get.email_column,
                'priority_column' : get.priority_column
            }
        except ObjectDoesNotExist:
            create = models.settings.objects.create(user=user)
            get = create
            dict={
                'start_row' : get.start_row,
                'gannt_start_column' : get.gannt_start_column,
                'week_number_row' : get.week_number_row,
                'start_column' : get.start_column,
                'end_column' : get.end_column,
                'pic_column' : get.pic_column,
                'email_column' : get.email_column,
                'priority_column' : get.priority_column
            }
        return render(request, 'settings.html',{
    'dict' : dict})
    else:
        get = models.settings.objects.get(user=user)
        get.gannt_start_column = request.POST['gannt_start_column']
        get.week_number_row = request.POST['week_number_row']
        get.start_row = request.POST['start_row']
        get.start_column = request.POST['start_column']
        get.end_column = request.POST['end_column']
        get.pic_column = request.POST['pic_column']
        get.email_column = request.POST['email_column']
        get.priority_column = request.POST['priority_column']
        get.save()
        return redirect('settings')
    
# Ini function untuk membaca file excel yang diupload pada melakukan new project
@login
def confirmation (request, id):
    if request.method == 'POST':
        # Mengambil settings akun yang login dari database
        get = models.settings.objects.get(user=request.user)

        allpriority = models.priority.objects.all()

        forprint = list()
        if 'upload' in request.POST:
            # LOAD XLSX
            link = request.POST['link']
            find = models.project.objects.get(id=id)
            find.link = link
            find.save()
            files = request.FILES['filea']
            wb = load_workbook(files)
            ws = wb.active
            
            # START GANTT CHART ROW
            start_row = get.start_row

            def get_date_range_for_week(start_year, end_year, week_start, week_end):
                if week_start is None or week_start < 1:
                    return []

                # Set semisal week_end sama dengan week_start
                if week_end is None:
                    week_end = week_start

                # Cari pertama di tahun itu
                def get_first_day_of_year(year):
                    first_day = datetime(year, 1, 1)
                    if first_day.weekday() != 0:  # Set ke hari senin
                        first_day -= timedelta(days=first_day.weekday())
                    return first_day

                # Cari untuk start_year dan end_year
                first_day_start_year = get_first_day_of_year(start_year)
                first_day_end_year = get_first_day_of_year(end_year)

                # Function mencari span start_date end_date senin-jumat
                def calculate_week_dates(first_day, week_number):
                    offset = (week_number - 1) * 7
                    start_date = first_day + timedelta(days=offset)
                    end_date = start_date + timedelta(days=4)  # Mencari hari jumat di week tersebut
                    return start_date, end_date

                date_ranges = []

                # Mencari Start dan End Date dari week given
                start_date, _ = calculate_week_dates(first_day_start_year, week_start)
                _, end_date = calculate_week_dates(first_day_end_year, week_end)
                date_ranges.append((start_date, end_date))

                return date_ranges
                
            def find_col_with_filled_color(ws, row_index):
                filled_cells = []
                filled_year = []
                year = None
                # START ITERATION FROM THE COLUMN GANTT_START_COLUMN STARTED
                for col_index in range(get.gannt_start_column, ws.max_column + 1):
                    cell = ws.cell(row=row_index, column=col_index)
                    
                    if isinstance(cell.fill.fgColor.theme, int) or (cell.fill.fgColor.rgb != 'FFFF0000' and  cell.fill.fgColor.rgb !='00000000' and cell.fill.fgColor.rgb !='FFFFFF00'):
                        findweek = ws.cell(get.week_number_row, col_index).value
                        findyear = ws.cell(get.week_number_row - 1, col_index).value
                        if '25' in str(findyear):
                            year = 2025
                        elif '26' in str(findyear):
                            year = 2026
                        else:
                            year = 2024
                        filled_year.append(year)
                        filled_cells.append(findweek)
                print(filled_cells)
                print(filled_year)
                return filled_cells, filled_year if filled_cells else None
            
            all = {}
            all['col_start'] = []
            for i in range(1, get.end_column - get.start_column):
                all[f'col{i}_values'] = []
            all['col_end'] = []

            # ITERATE COLUMN START_ROW AS A BASE
            for row in ws.iter_rows(min_row=start_row,min_col=get.start_column,max_col=get.start_column, values_only=True):
                cell_value = row[0] if row and row[0] is not None else None
                all['col_start'].append(cell_value)  
            # ITERATE ROW FOR EVERY COLUMN TO GET THE VALUES
            for i in range(1, get.end_column - get.start_column):
                for row in ws.iter_rows(min_row=start_row, min_col=get.start_column+i, max_col=get.start_column+i, values_only=True):
                    cell_value = row[0] if row and row[0] is not None else None
                    all[f'col{i}_values'].append(cell_value)
            for row in ws.iter_rows(min_row=start_row, min_col=get.end_column, max_col=get.end_column, values_only=True):
                cell_value = row[0] if row and row[0] is not None else None
                all['col_end'].append(cell_value)        
            
            
            tasks_data = []
            current_task = None
        # Dictionary to keep track of the current parent task at each column level
            current_tasks = {key: None for key in all.keys()}

            for idx, value in enumerate(all['col_start']):
                if value is not None:
                    # Main Task
                    p, year = find_col_with_filled_color(ws, idx + start_row)
                    if p is not None and len(p) > 0:
                        if len(p) > 1:
                            week_start = min(p)
                            week_end = max(p)
                            if p[0] > p[-1]:
                                week_end = p[-1]
                                week_start = p[0]
                            start_year = min(year)
                            end_year = max(year)
                            if year:
                                get_ranges = get_date_range_for_week(start_year,end_year, week_start, week_end)
                                for start_date, end_date in get_ranges:
                                    start_date = start_date.strftime('%Y-%m-%d')
                                    end_date = end_date.strftime('%Y-%m-%d')
                        else:
                            print('elses')
                            print(p)
                            week_start = min(p)
                            week_end = None
                            start_year = min(year)
                            end_year = max(year)
                            if year:
                                get_ranges = get_date_range_for_week(start_year,end_year, week_start, week_end)
                                for start_date, end_date in get_ranges:
                                    start_date = start_date.strftime('%Y-%m-%d')
                                    end_date = end_date.strftime('%Y-%m-%d')
                        # Get Ranges masih manual pakai tahun 2024
                    else:
                        week_start = None
                        week_end = None
                        start_date = week_start
                        end_date = week_end

                    findpic = ws.cell(idx + start_row, get.pic_column).value
                    if findpic is not None:
                        findpic = findpic.upper()
                    findperson = ws.cell(idx + start_row, get.email_column).value
                    if findperson is not None:
                        if ',' in findperson:
                            findperson = findperson.split(',')
                        else:
                            findperson = [findperson]
                    findpriority = ws.cell(idx + start_row, get.priority_column).value
                    if findpriority is not None:
                        findpriority = findpriority.capitalize()

                    current_task = {
                        'name': value,
                        'start_date': start_date,
                        'due_date': end_date,
                        'pic': findpic,
                        'person': findperson,
                        'priority' : findpriority,
                        'subtasks': []
                    }
                    tasks_data.append(current_task)

                    # Set the current task for col_start level
                    current_tasks['col_start'] = current_task

                    # Reset the current task for all subsequent columns
                    for key in list(current_tasks.keys())[1:]:
                        current_tasks[key] = None
                    print(value)
                    print('start:', start_date)
                    print('due:', end_date)


                else:
                    # Iterasi kembali pada column yang tersisa untuk mencari subtasks
                    for key in list(all.keys())[1:]:
                        col_values = all[key]
                        subtask_value = col_values[idx]
                        if subtask_value is not None:
                            p, year = find_col_with_filled_color(ws, idx + start_row)
                            if p is not None and len(p) > 0:
                                if len(p) > 1:
                                    week_start = min(p)
                                    week_end = max(p)
                                    print(p)
                                    if p[0] > p[-1]:
                                        week_end = p[-1]
                                        week_start = p[0]
                                    start_year = min(year)
                                    end_year = max(year)
                                    if year:
                                        get_ranges = get_date_range_for_week(start_year,end_year, week_start, week_end)
                                        for start_date, end_date in get_ranges:
                                            start_date = start_date.strftime('%Y-%m-%d')
                                            end_date = end_date.strftime('%Y-%m-%d')
                                else:
                                    print('elses')
                                    print(p)
                                    week_start = min(p)
                                    week_end = None
                                    start_year = min(year)
                                    end_year = max(year)
                                    if year:
                                        get_ranges = get_date_range_for_week(start_year,end_year, week_start, week_end)
                                        for start_date, end_date in get_ranges:
                                            start_date = start_date.strftime('%Y-%m-%d')
                                            end_date = end_date.strftime('%Y-%m-%d')
                            else:
                                week_start = None
                                week_end = None
                                start_date = week_start
                                end_date = week_end

                            findpic = ws.cell(idx + start_row, get.pic_column).value
                            if findpic is not None:
                                findpic = findpic.upper()
                            findperson = ws.cell(idx + start_row, get.email_column).value
                            if findperson is not None:
                                if ',' in findperson:
                                    findperson = findperson.split(',')
                                else:
                                    findperson = [findperson]
                            findpriority = ws.cell(idx + start_row, get.priority_column).value
                            if findpriority is not None:
                                findpriority = findpriority.capitalize()

                            subtask = {
                                'name': subtask_value,
                                'pic': findpic,
                                'start_date': start_date,
                                'due_date': end_date,
                                'person': findperson,
                                'priority' : findpriority,
                                'subtasks': []
                            }

                            # print(subtask_value)
                            # print('start:', start_date)
                            # print('due:', end_date)
                            # Memberikan key parent untuk mengetahui parent tasks
                            for parent_key in reversed(list(current_tasks.keys())[:list(current_tasks.keys()).index(key)]):
                                if current_tasks[parent_key] is not None:
                                    if 'subtasks' not in current_tasks[parent_key]:
                                        current_tasks[parent_key]['subtasks'] = []
                                    current_tasks[parent_key]['subtasks'].append(subtask)
                                    break

                            # Set the current task for the current column
                            current_tasks[key] = subtask

                            # Reset the current task for all subsequent columns
                            for subsequent_key in list(current_tasks.keys())[list(current_tasks.keys()).index(key) + 1:]:
                                current_tasks[subsequent_key] = None

            
            # IT IS USED TO ATTACH A PARENT NAME KEYS FOR A SUBTASKS BASED ON A SUBTASKS THAT CONTAINED IN PARENT
            def process_tasks(tasks, parent_hierarchy=None):
                if parent_hierarchy is None:
                    parent_hierarchy = []
                    
                for task in tasks:
                    task['Parent'] = parent_hierarchy  # Assign the current parent hierarchy list to 'Parent' key
                    forprint.append(task)
                    
                    # Recursively process subtasks
                    if 'subtasks' in task and task['subtasks']:
                        # Pass the updated hierarchy, including the current task's name, to the subtasks
                        process_tasks(task['subtasks'], parent_hierarchy + [task['name']])
                        
                    # Remove 'subtasks' key to flatten the structure
                    if 'subtasks' in task:
                        del task['subtasks']
            
            
            process_tasks(tasks_data)
            print(forprint)
            
            
            # SESSION THE DATA TO BE PASSED INTO AFTER POST LOGIC

            request.session['forprint'] = forprint
            request.session['name'] = id
            
            return render (request, 'confirmation.html', {
                'tasks_data' : forprint,
                'allpriority' : allpriority
            })
        else:
            # IN THIS SECTION, THE DATA IS BEING CREATED INTO REDMINE DB
            # GET THE SESSION
            forprint = request.session.get('forprint',[])
            name = request.session.get('name',[])
            def create_issue(item):
                # THE ['name'] IS USE TO HANDLE A MULTIPLE FORM FOR A MULTIPLE TASK THAT WILL BE PROCCESSED
                
                form_name = f"name_{item['name']}"
                form_start_date = f"start_date_{item['name']}"
                form_due_date = f"due_date_{item['name']}"
                form_responsible = f"responsible_{item['name']}[]"
                form_priority = f"priority_{item['name']}"
                form_parent = f"parenttask_{item['name']}"
                subject = request.POST.get(form_name)
                start_date = request.POST.get(form_start_date)
                due_date = request.POST.get(form_due_date)
                assigned_to = request.user
                parent = request.POST.get(form_parent)
                priority = request.POST.get(form_priority)
                
                responsible = request.POST.getlist(form_responsible)
                ids = models.project.objects.get(id=id)
                pri = models.priority.objects.get(id=priority)
                start_date = parse_date(start_date) if start_date else None
                due_date = parse_date(due_date) if due_date else None
                print(parent)
                create = models.task(
                    id_project = ids,
                    subject = subject, 
                    assignee = assigned_to,
                    start_date = start_date,
                    due_date = due_date,
                    id_priority = pri,)
                if parent:
                    create.parent = parent
                create.save()
                
                distinct = list(set(responsible))
                for email in distinct:
                    getid = models.task.objects.get(id=create.id)
                    pic = models.pic(id_task=getid, pic=email)
                    pic.save()
            for item in forprint:
                create_issue(item)
            
            return redirect('listproject')
    return render(request, 'confirmation.html')

@login
def deleteproject(request,id):
    get = models.project.objects.get(id=id)
    get.delete()
    return redirect('listproject')

@login
def updateproject(request, id):
    update = models.project.objects.get(id=id)
    getall = models.status.objects.exclude(id=2)
    print(getall)
    if request.method == "GET":
        return render(request, 'updateproject.html',{
            'update' : update,
            'getall' : getall
        })
    else:
        subject = request.POST.get('subject', '')  # Safely retrieve 'subject' from POST data
        description = request.POST.get('description', '')  # Safely retrieve 'description' from POST data
        link = request.POST.get('link', '')
        status = request.POST['status']
        getstatus = models.status.objects.get(id=status)
        # Update the project object with the retrieved values
        update.status = getstatus
        update.subject = subject
        update.desc = description
        update.link = link

        update.save()
        return redirect ('listproject')


@login
def listissue(request,id):
    listissue = models.task.objects.filter(id_project = id)
    id = request.session['idproject'] = id
    return render(request, 'listissue.html', {
        'listissue' : listissue,
        'id' : id
        
    })

def listdetails(request, id):
    if request.method == 'GET':
        get = models.task.objects.get(id=id)
        getproject = get.id_project
        getpic = models.pic.objects.filter(id_task = id)
        if get.parent:
            if "[" in get.parent:
                parent = ast.literal_eval(get.parent)
                parent = [s for s in parent if 'phase' not in s.lower()]
            else:
                parent = [get.parent]
        else:
            parent = None
        listpic = []
        for item in getpic:
            listpic.append(item.pic)
        listpic = ','.join(listpic)
        email = request.GET.get('email', '')
        if email:
            email = unquote(email)
        return render(request, 'listdetails.html', {
        "get" : get,
        "getproject" : getproject,
        'getpic' : listpic,
        'email' : email,
        'parent' : parent,
        'id' : id
    })
    else:
        sender = request.POST['sender']
        description = request.POST['description']
        get = models.task.objects.get(id=id)
        receipent = get.assignee.email
        getstatus = models.status.objects.get(jenis = "Under Review")
        get.status = getstatus
        fullname = get.assignee.first_name + " " + get.assignee.last_name
        email_api = "http://10.24.7.70:3333/send-email"
        subject = f"#{get.id} [{get.subject}] - {get.id_project} Task Feedback"
        body = f"""
                Dear {fullname},

                This is my feedback about #{get.id} [{get.subject}] task from {get.id_project}:
                {description}
                
                Regards,
                {sender}
                """
        payload = {
            "to": [receipent],
            "cc": [],
            "subject": subject,
            "body": body
        }
        print(payload)
        print(datetime.now())
        response = requests.post(email_api, json=payload)
        if response.status_code == 200:
            print("Email sent successfully.")
            get.save()
        else:
            print(f"Failed to send email. Status code: {response.status_code}")
            print(response.text)
        return redirect ('listdetails', id=id)
        #  
    

@login
def newissue(request):
    allissue = models.task.objects.filter(assignee = request.user)
    
    return render(request, 'newissue.html',{
        'allissue' : allissue
    })

@login
def updateissue(request,id):
    get = models.task.objects.get(id=id)
    allpriority = models.priority.objects.all()
    getpic = models.pic.objects.filter(id_task = id)
    allstatus = models.status.objects.all()
    if request.method == "GET":
        
        return render(request, 'updateissue.html', {
            'update' : get,
            'allpriority' : allpriority,
            'getpic' : getpic,
            'allstatus' : allstatus
        })
    else:
        subject = request.POST['subject']
        status = request.POST['status']
        start_date = request.POST['start_date']
        due_date = request.POST['due_date']
        priority = request.POST['priority']
        addpic = request.POST['addpic']
        pic = request.POST.getlist("pic")
        parent = request.POST['parent']

        get.subject = subject
        getstatus = models.status.objects.get(id=status)
        getpriority = models.priority.objects.get(id=priority)
        get.status = getstatus
        
        gettask = models.task.objects.get(id=id)
        findpic = models.pic.objects.filter(id_task = id)
        findpic.delete()
        
        for i in pic:
            newpic = models.pic.objects.create(id_task = gettask, pic=i)
            newpic.save()
        if start_date:
            get.start_date = start_date
        if due_date:
            get.due_date = due_date
        get.parent = parent
        get.id_priority = getpriority
      
        gettask = models.task.objects.get(id=id)
        if addpic:
            registerpic = models.pic.objects.create(id_task=gettask, pic=addpic)
            registerpic.save()
        get.save()
        return redirect('listdetails', id=id)

@login
def closeissue(request, id):
    get = models.task.objects.get(id=id)
    if get.status.id == 3:
        getstatus = models.status.objects.get(id=1) #Get open status
    else:
        getstatus = models.status.objects.get(id=3) #Get close status
    get.status = getstatus
    get.save()
    return redirect('listdetails', id=get.id)
@login
def deleteissue(request,id):
    get = models.task.objects.get(id=id)
    get.delete()
    
    return redirect('listissue', id=get.id_project.id)

def user_login(request):
    if request.method == "GET":
        # FIND SESSION FROM RECENT LOGIN
        if request.session.get('username') and request.session.get('password'):
            return redirect('index')
        else:
            return render(request, 'login.html')
    else:
        username = request.POST['username']
        password = request.POST['password']
        userobj = authenticate(request, username=username, password=password)
        if userobj is not None :
            auth_login(request, userobj)
            request.session['username'] = username
            request.session['password'] = password
            return redirect("index")
        else:
            messages.error(request, "Username atau Password Anda Salah!")
            return redirect("login")
@login
def user_logout (request):
    logout(request)
    return redirect ('login')

@login
def send_email(request):
    if request.method == "GET":
        today = date.today()
        
        # Query tasks
        getstatusopen = models.status.objects.get(id=1)
        start_today_tasks = models.task.objects.filter(start_date=today,status=getstatusopen).exclude(due_date=today)

# Tasks that are due today
        end_today_tasks = models.task.objects.filter(
            due_date=today,
            status=getstatusopen
        ).exclude(start_date=today)

        # Ongoing tasks excluding tasks that start and end today
        ongoing_tasks = models.task.objects.filter(
            assignee=request.user,
            start_date__lt=today,  # Started before today
            due_date__gt=today,    # Due after today
            status=getstatusopen
        )
                

        def find_pic(task):
            return list(models.pic.objects.filter(id_task=task))

        def create_email_body(task, status, recipient_email):
            general_name = "Team Member"
            fullname = task.assignee.first_name + " " + task.assignee.last_name
            author_name = fullname            
            project_name = task.id_project.subject
            subject = task.subject
            start_date = task.start_date
            due_date = task.due_date
            task_id = task.id
            task_parent = task.parent
            task_link = task.id_project.link
            encoded_email = urllib.parse.quote(recipient_email)
            if status == 'start':
                body = f"""
                Dear {general_name},

                You have a task that started today and will be due on {due_date} from the {project_name} project
                with the task subject: {subject} of {task_parent}.

                Give your feedback at: http://10.24.7.165/listdetails/{task_id}?email={encoded_email}
                
                or view our project timeline at {task_link}
                
                Regards,
                {author_name}
                """
            elif status == 'ongoing':
                body = f"""
                Dear {general_name},

                You still have a running task that started on {start_date} and will be due on {due_date}
                from the {project_name} project with the task subject: {subject} of {task_parent}.

                Give your feedback at: http://10.24.7.165/listdetails/{task_id}?email={encoded_email}
                
                or view our project timeline at {task_link}
                
                Regards,
                {author_name}
                """
            else:
                body = f"""
                Dear {general_name},

                You have a task that will be due today ({due_date}) from the {project_name} project
                with the task subject: {subject} of {task_parent}.

                Give your feedback at: http://10.24.7.165/listdetails/{task_id}?email={encoded_email}

                or view our project timeline at {task_link}                

                Regards,
                {author_name}
                """
            return body

        def send_email(to, cc, subject, body):
            email_api = "http://10.24.7.70:3333/send-email"
            payload = {
                "to": [to],
                "cc": [cc],
                "subject": subject,
                "body": body
            }
            print(payload)
            response = requests.post(email_api, json=payload)
            if response.status_code == 200:
                print("Email sent successfully.")
            else:
                print(f"Failed to send email. Status code: {response.status_code}")
                print(response.text)

        # # Process tasks and send emails
        for task in start_today_tasks:
            pics = find_pic(task)
            for pic in pics:
                body = create_email_body(task, 'start', pic.pic)
                send_email(pic.pic, task.assignee.email, f"#{task.id} [{task.subject}] Task Reminder", body)
        
        for task in end_today_tasks:
            pics = find_pic(task)
            for pic in pics:
                body = create_email_body(task, 'end', pic.pic)
                send_email(pic.pic, task.assignee.email, f"#{task.id} [{task.subject}] Task Reminder", body)
        
        for task in ongoing_tasks:
            pics = find_pic(task)
            for pic in pics:
                body = create_email_body(task, 'ongoing', pic.pic)
                send_email(pic.pic, task.assignee.email, f"#{task.id} [{task.subject}] Task Reminder", body)
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        return redirect('index')
    
@login
def register(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        email = request.POST['email']
        create = models.User.objects.create_user(username=username, email=email, password=password, first_name=first_name, last_name=last_name)
        create.save()
        return redirect ('register')
    return render(request, 'register.html',{

    })

@login
def export_csv(request, id):
    
    getproject = models.project.objects.get(id=id)
    get = models.task.objects.filter(id_project = getproject)
    getpic = models.pic.objects.filter(id_task__in = get)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{getproject.subject}.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Task', 'Start Date', 'Due Date', 'PIC', 'Status', 'Priority'])  # Add your actual column headers here

    for task in get:
        task_pics = getpic.filter(id_task=task)
        for pic in task_pics:
            writer.writerow([
                task.id,
                task.subject,  # Replace with actual task name field
                task.start_date,  # Replace with actual start date field
                task.due_date,  # Replace with actual due date field
                pic.pic,  # Replace with actual pic name field
                task.status,  # Replace with actual status field
                task.id_priority  # Replace with actual priority field
            ])

    return response

def handler404(request, exception, template_name='404.html'):
    response = render(template_name)
    response.status_code = 404
    return response

def handler500(request,template_name='500.html'):
    response = render(template_name)
    response.status_code = 500
    return response